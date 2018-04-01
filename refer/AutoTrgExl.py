#!/usr/bin/env python
import os
import urllib2
import urllib
import re
import sys
import subprocess
import time
import json
import optparse
import copy
import ssl
from openpyxl import Workbook
from openpyxl import load_workbook


def GetRunConfigs(uuid):
    eris_url = "https://eris-portal.nvidia.com/OneSubmissionTestSuitesServlet?uuid=%s&tableselector=dataTableOneSubmissionBuilds&url"%uuid
    eris_results = GetURLSource(eris_url)
    AllRunConfigs= re.findall(r"badge-(failed|aborted)\'>(.*?)<\\/span>.*?historyChannelLabel.*?testSuiteNameText=(.*?)&.*?ConfigID=(.*?)\'>",eris_results)
    configID_list = list()
    for i in AllRunConfigs:
        configID_list.append(i[3])
    configID_list = list(set(configID_list))
    configID_dict = dict()
    for i in configID_list:
        configID_dict[i]=ConfigID(i)
    Configs= list()
    for i in AllRunConfigs:
        config = list()
        lists = list(i)
        lists.extend(configID_dict[i[3]])
        config = [lists[2],lists[3],lists[6],lists[8],lists[7],lists[11],lists[1]]
        if config[5] == "[Default]": config[5] = ""
        if config[6] != "Failed Other":
            Configs.append(config)
    return Configs

def ConfigID(cid):
    url = "https://eris-portal.nvidia.com/GetTestHeaderServlet?testSuiteNameText=&testNameText=&erisConfigID=%s"%cid
    url_source = GetURLSource(url)
    res_td = r"<td>(.*?)<\\/td>"
    config = re.findall(res_td,url_source)
    return config

def Getbuildresults(uuid,Components):
    eris_url="https://eris-portal.nvidia.com/OneSubmissionBuildsServlet?uuid=%s&tableselector=dataTableOneSubmissionTestSuites&url"%uuid
    eris_results = GetURLSource(eris_url)
    Allbuilds=re.findall(r"badge-(failed|aborted)\'>(.*?)<\\/span>.*?&componentNameText=(.*?)&.*?ConfigID=(.*?)\'>",eris_results)
    configID_list = list()
    for i in Allbuilds:
        configID_list.append(i[3])
    configID_list = list(set(configID_list))
    configID_dict = dict()
    for i in configID_list:
        configID_dict[i]=ConfigID(i)
    buildresults = list()
    for build in Allbuilds:
        config  = list()
        lists = list(build)
        lists.extend(configID_dict[build[3]])
        #lists = ['failed','1','cuda_apps_tests_L0','3159','', '', '', 'x86_64', 'FC23', '[Default]', 'release', 'pgi16_7']
        config = [lists[2],lists[3],lists[8],lists[7],lists[11],lists[1]]
        if config[4] == "[Default]": config[4] = ""
        buildresults.append(config)
    return buildresults

def SelectTestsuites(Allstatus,Testsuites):
    Selected = []
    for items in Allstatus:
        if items[0] in Testsuites: Selected.append(items)
    return Selected

def SelectFailTestsuites(Allstatus,Testsuites,select):
    if select:
        SelectedSuites = SelectTestbyoption(Allstatus,Testsuites)
    else:
        SelectedSuites = SelectTestsuites(Allstatus,Testsuites)
    return SelectedSuites

def Selectnewcomplete(testsuites,uuid):
    uuid_dir = os.path.join(sys.path[0],"uuid_file")
    if not os.path.exists(uuid_dir):
        os.mkdir(uuid_dir)
    uuid_file = os.path.join(uuid_dir,'%s.out'%uuid)
    if os.path.isfile(uuid_file):
        old_testsuites = testsuites
    else: old_testsuites = list()
    new_testsuites = list()
    for i in testsuites:
        with open(uuid_file,'a+') as file:
            if ','.join(i)+'\n' not in file.readlines():
                file.write(','.join(i) + '\n')
                new_testsuites.append(i)
            else: pass
    return old_testsuites,new_testsuites

def GetSubInfo(uuid):
    url = 'https://eris-portal.nvidia.com/GetSubmissionHeaderOneUuidServlet?uuid=%s'%uuid
    header_source = GetURLSource(url)
    prodcut = re.findall('<td>(.*?)<',header_source)[0]
    tag = re.findall('<td>(.*?)<',header_source)[5]
    subtime = re.findall('DoOneSubmissionViewCommand.*?<td>(.*?)<',header_source)[0].split(' ')[0]
    vlcp = GetProduct(uuid)
    return prodcut,tag,subtime,vlcp

def GetProduct(uuid):
    url = 'https://eris-portal.nvidia.com/GetProductServlet?uuid=%s'%uuid
    headersource = GetURLSource(url)
    config=json.loads(headersource)
    product = config["product"]
    vlcps = {
                      "CUDA Development"               : "//sw/gpgpu/eris/cuda.vlcp",
                      "CUDA 7.0 r346"                  : "//sw/rel/gpgpu/toolkit/r7.0/eris/cuda70.vlcp",
                      "CUDA 8.0 r361"                  : "//sw/rel/gpgpu/toolkit/r8.0/eris/cuda.vlcp",
                      "CUDA 8.0"                       : "//sw/rel/gpgpu/toolkit/r8.0/eris/cuda.vlcp",
                      "CUDA 8.0 r367"                  : "//sw/rel/gpgpu/toolkit/r8.0/eris/cuda_r367.vlcp",
                      "CUDA 8.0 r375"                  : "//sw/rel/gpgpu/toolkit/r8.0/eris/cuda_r375.vlcp",
                      "CUDA 8.0 chips_a"               : "//sw/rel/gpgpu/toolkit/r8.0/eris/cuda_chips_a.vlcp",
                      "CUDA 8.0 r375 mobile"           : "//sw/rel/gpgpu/toolkit/r8.0/eris/cuda_r375_mobile.vlcp",
                      "CUDA gpgpu cuda_a"              : "//sw/gpgpu/eris/cuda.vlcp",
                      "CUDNN V6 CUDA 8.0 r375"       : "//sw/gpgpu/MachineLearning/cudnn_v6/eris/cudnn_r80_r375.vlcp",
                      "CUDNN V6 CUDA 7.5 r352"       : "//sw/gpgpu/MachineLearning/cudnn_v6/eris/cudnn_75.vlcp",
                      "CUDNN V6 CUDA 8.0 r361"       : "//sw/gpgpu/MachineLearning/cudnn_v6/eris/cudnn_80.vlcp",
                      "CUDNN V6 CUDA 8.0 r367"       : "//sw/gpgpu/MachineLearning/cudnn_v6/eris/cudnn_r80_r367.vlcp",
					  "CUDNN V7 CUDA 9.0 r384"       : "//sw/gpgpu/MachineLearning/cudnn_v7/eris/cudnn_r90_r384.vlcp",
					  "CUDNN V7 CUDA 9.1 r387"       : "//sw/gpgpu/MachineLearning/cudnn_v7/eris/cudnn_r91_r387.vlcp",
					  "CUDNN V7 CUDA 8.0 r375"       : "//sw/gpgpu/MachineLearning/cudnn_v7/eris/cudnn_r80_r375.vlcp",
                      "CUDNN DEV GPGPU CUDA_A"       : "//sw/gpgpu/MachineLearning/cudnn/eris/cudnn_gpgpu_cuda_a.vlcp",
                      "CUDNN DEV CUDA 9.1 r387"       : "//sw/gpgpu/MachineLearning/cudnn/eris/cudnn_r91_r387.vlcp",
					  "CUDNN V7.1 CUDA 9.0 r384"       : "//sw/gpgpu/MachineLearning/cudnn_v7.1/eris/cudnn_r90_r384.vlcp",
  					  "CUDNN V7.1 CUDA 8.0 r375"       : "//sw/gpgpu/MachineLearning/cudnn_v7.1/eris/cudnn_r80_r375.vlcp",
					  "CUDNN V7.1 CUDA 9.1 r387"       : "//sw/gpgpu/MachineLearning/cudnn_v7.1/eris/cudnn_r91_r387.vlcp",
					  "CUDNN V7.1 GPGPU CUDA_A"       : "//sw/gpgpu/MachineLearning/cudnn_v7.1/eris/cudnn_gpgpu_cuda_a.vlcp",
                      "GIE V1.0 CUDA 7.5 r352"         : "//sw/gpgpu/MachineLearning/DIT/release/1.0/eris/gie_75.vlcp",
                      "GIE V1.0 CUDA 8.0 r361"         : "//sw/gpgpu/MachineLearning/DIT/release/1.0/eris/gie_80.vlcp",
                      "GIE V1.0 CUDA 8.0 r367"         : "//sw/gpgpu/MachineLearning/DIT/release/1.0/eris/gie_80_r367.vlcp",
                      "GIE V1.0 CUDA 8.0 r375 mobile"  : "//sw/gpgpu/MachineLearning/DIT/release/1.0/eris/gie_80_r375_mobile.vlcp",
                      "GIE V1.0 CUDA 7.5 r352"         : "//sw/gpgpu/MachineLearning/DIT/release/1.0/eris/gie_75.vlcp",
                      "GIE V1.0 CUDA 8.0 r361"         : "//sw/gpgpu/MachineLearning/DIT/release/1.0/eris/gie_80.vlcp",
                      "GIE V1.0 CUDA 8.0 r367"         : "//sw/gpgpu/MachineLearning/DIT/release/1.0/eris/gie_80_r367.vlcp",
                      "GIE V1.0 CUDA 8.0 r375"         : "//sw/gpgpu/MachineLearning/DIT/release/1.0/eris/gie_80_r375.vlcp",
                      "GIE V2.0 CUDA 8.0 r375 mobile"  : "//sw/gpgpu/MachineLearning/DIT/release/2.0/eris/gie_80_r375_mobile.vlcp",
                      "GIE V2.1 CUDA 7.5 r352"         : "//sw/gpgpu/MachineLearning/DIT/release/2.1/eris/gie_75.vlcp",
                      "GIE V2.0 CUDA 8.0 r361"         : "//sw/gpgpu/MachineLearning/DIT/release/2.0/eris/gie_80.vlcp",
                      "GIE V2.0 CUDA 8.0 r367"         : "//sw/gpgpu/MachineLearning/DIT/release/2.0/eris/gie_80_r367.vlcp",
                      "GIE V2.1 CUDA 8.0 r375"         : "//sw/gpgpu/MachineLearning/DIT/release/2.1/eris/gie_80_r375.vlcp",
                      "GIE Trunk CUDA 7.5 r352"        : "//sw/gpgpu/MachineLearning/DIT/trunk/eris/gie_75.vlcp",
                      "GIE Trunk CUDA 8.0 r361"        : "//sw/gpgpu/MachineLearning/DIT/trunk/eris/gie_80.vlcp",
                      "GIE Trunk CUDA 8.0 r367"        : "//sw/gpgpu/MachineLearning/DIT/trunk/eris/gie_80_r367.vlcp",
                      "GIE Trunk CUDA 8.0 r375"        : "//sw/gpgpu/MachineLearning/DIT/trunk/eris/gie_80_r375.vlcp",
                      "Heterogeneous OpenMP Runtime gpgpu cuda_a"   : "//sw/gpgpu/runtimes/omp/omp.vlcp",
                      "ICE 1.1 CUDA 7.5 r352"          : "//sw/gpgpu/eris/ice11.vlcp",
                      "CAFFE CUDNN V6 NCCL Private V1.6 CUDA 8.0 r361" : "//sw/gpgpu/MachineLearning/caffe/eris/caffetest_cudnnV6_nccl_privateV1.6_80_r361.vlcp",
                      "CAFFE CUDNN V6 NCCL Public V1.3 CUDA 8.0 r361" : "//sw/gpgpu/MachineLearning/caffe/eris/caffetest_cudnnV6_nccl_publicV1.3_80_r361.vlcp",
                      "CAFFE CUDNN V6 NCCL Private V1.6 CUDA 8.0 r367" : "//sw/gpgpu/MachineLearning/caffe/eris/caffetest_cudnnV6_nccl_privateV1.6_80_r367.vlcp",
                      "CAFFE CUDNN V6 NCCL Public V1.3 CUDA 8.0 r367" : "//sw/gpgpu/MachineLearning/caffe/eris/caffetest_cudnnV6_nccl_publicV1.3_80_r367.vlcp",
                      "CAFFE CUDNN V6 NCCL Private V1.6 CUDA 8.0 r375" : "//sw/gpgpu/MachineLearning/caffe/eris/caffetest_cudnnV6_nccl_privateV1.6_80_r375.vlcp",
                      "CAFFE CUDNN V6 NCCL Public V1.3 CUDA 8.0 r375" : "//sw/gpgpu/MachineLearning/caffe/eris/caffetest_cudnnV6_nccl_publicV1.3_80_r375.vlcp",
                      "CAFFE V1.6 CUDNN V7 CUDA 9.0 r384" : "//sw/gpgpu/MachineLearning/caffe/eris/caffetest_cudnnV7_nccl_V1.6_90_r384.vlcp",
                      "OpenCL gpgpu r375"              : "//sw/gpgpu/opencl/eris/r375_00/opencl.vlcp",
                      "OpenCL gpgpu r367"              : "//sw/gpgpu/opencl/eris/r367_00/opencl.vlcp",
                      "OpenCL gpgpu cuda_a"            : "//sw/gpgpu/opencl/eris/cuda_a/opencl.vlcp",
                      "OpenMPI gpgpu cuda_a"           : "//sw/gpgpu/eris/openmpi.vlcp",
                      "OpenMPI CUDA 8.0 r361"          : "//sw/rel/gpgpu/toolkit/r8.0/eris/openmpi.vlcp",
                      "OpenMPI CUDA 8.0 r367"          : "//sw/rel/gpgpu/toolkit/r8.0/eris/openmpi_r367_gp.vlcp",
                      "CAFFE CUDNN V5.1 CUDA 7.5 r352" : "//sw/gpgpu/MachineLearning/caffe/eris/caffe_cudnnV5.1_75.vlcp",
                      "CAFFE CUDNN V5.1 CUDA 8.0 r361" : "//sw/gpgpu/MachineLearning/caffe/eris/caffe_cudnnV5.1_80.vlcp",
                      "CAFFE CUDNN V5.1 CUDA 8.0 r367" : "//sw/gpgpu/MachineLearning/caffe/eris/caffe_cudnnV5.1_r80_r367.vlcp",
                      "CAFFE CUDNN V6 CUDA 7.5 r352"   : "//sw/gpgpu/MachineLearning/caffe/eris/caffe_cudnnV6_75.vlcp",
                      "CAFFE CUDNN V6 CUDA 8.0 r361"   : "//sw/gpgpu/MachineLearning/caffe/eris/caffe_cudnnV6_80.vlcp",
                      "CAFFE CUDNN V6 CUDA 8.0 r367"   : "//sw/gpgpu/MachineLearning/caffe/eris/caffe_cudnnV6_r80_r367.vlcp",
                      "DIGITS Development"             : "//sw/gpgpu/MachineLearning/digits/eris/digits-dev.vlcp",
                      "UVM gpgpu cuda_a"               : "//sw/gpgpu/eris/uvm.vlcp",
                      "UVM 8.0 chips_a"                : "//sw/rel/gpgpu/toolkit/r8.0/eris/uvm_chips_a.vlcp",
                      "UVM 8.0 r361"                   : "//sw/rel/gpgpu/toolkit/r8.0/eris/uvm_r361.vlcp",
                      "UVM 8.0 r367"                   : "//sw/rel/gpgpu/toolkit/r8.0/eris/uvm_r367.vlcp",
                      "UVM 8.0 r375"                   : "//sw/rel/gpgpu/toolkit/r8.0/eris/uvm_r375.vlcp"
                    }
    return vlcps[product]

def Addtestcmd(uuid,testsuites,vlcp):
    for i in testsuites:
        if i is not None:
            cmdlist = list()
            if i[5]:
                compiler="--target-compiler=%s"%i[5]
            else: compiler=""
            if '32on64' in i[0]:
                test_arch="--test-system-arch=x86_64"
            else: test_arch=""
            tags = "Test_%s_%s_%s_%s%s"%(i[0],i[2],i[3],i[4],i[5])
            cmd = "vulcan --keep-going -v --eris --product=%s --build --testsuite %s --target-revision=cl-tot --target-gpu=%s --target-os=%s --target-arch=%s %s %s --tag=%s"%(vlcp,i[0],i[2],i[3],i[4],test_arch,compiler,tags)
            cmdlist = [cmd]
            i.extend(cmdlist)
    return testsuites

def Addbuildcmd(uuid,buildsuites,vlcp):
    for i in buildsuites:
        if i is not None:
            cmdlist = list()
            if i[4]:
                compiler="--target-compiler=%s"%i[4]
            else: compiler=""
            if i[2] == "agnostic":
                os = "Linux"
            else: os = i[2]
            tags = "Build_%s__%s_%s%s"%(i[0],i[2],i[3],i[4])
            cmd = "vulcan --keep-going -v --eris --product=%s --build %s --target-revision=cl-tot  --target-os=%s --target-arch=%s %s --tag=%s"%(vlcp,i[0],os,i[3],compiler,tags)
            cmdlist = [cmd]
            i.extend(cmdlist)
    return buildsuites

def Addtestlog(uuid,suites,info):
    for suite in suites:
        url = "https://eris-portal.nvidia.com/TestSuiteDetailsDataTableServlet?&sSearch_28=%s5E%s%s24&sSearch_29=%sE%s%s24&sSearch_30=%s5E%s%s24&sSearch_33=%s5E%s%s24&productNameText=%s&uuid1=%s&testSuiteNameText=%s&command=list&attr=fail"%('%',suite[2].replace('+','%5C%2B'),'%','%',suite[4],'%','%',suite[3],'%','%',suite[5],'%',info[0].replace(' ','+'),uuid,suite[0])
        try:
            url_source = GetURLSource(url)
            dict_obj = json.loads(url_source)
            test_source = dict_obj["jsonArray"][0][1]
            logurlparts = re.findall(r'&hostName=(.*?)&phase=(.*?)\"',test_source)
            logurllist = logurlparts[0]
        except:
            suite.extend(['','','Failed getting log'])
            continue
        if not 'pending_placeholder' in logurllist[0]:
            logurl = 'https://eris-portal.nvidia.com/secure/portal.jsp?&l.u=http://eris-fs000/logs/%s/%s_%s.log&t.t=t&tab=tabLV'%(uuid,logurllist[0],logurllist[1])
            suite.extend(['','',logurl])
        else: 
            suite.extend(['','','No log file'])
    return suites

def Addbuildlog(uuid,builds,info):
    for build in builds:
        url = "https://eris-portal.nvidia.com/BuildResultsListDataTableServlet?iDisplayStart=0&iDisplayLength=10&iSortCol_0=3&sSortDir_0=asc&uuid1=%s&productNameText=%s&submTag=%s&componentNameText=%s&configID=%s&command=list&attr=fail"%(uuid,info[0].replace(' ','+'),info[1],build[0],build[1])
        try: 
            url_source = GetURLSource(url)
            dict_obj = json.loads(url_source)
            build_source = dict_obj["jsonArray"][0][2]
            logurlparts = re.findall(r'&hostName=(.*?)&phase=(.*?)>',build_source)
            logurllist = logurlparts[0]
        except:
            build.extend(['','','Failed getting log'])
            continue
        if not 'pending_placeholder' in logurllist[0]:
            logurl = 'https://eris-portal.nvidia.com/secure/portal.jsp?&l.u=http://eris-fs000/logs/%s/%s_%s.log&t.t=b&tab=tabLV'%(uuid,logurllist[0],logurllist[1])
            build.extend(['','',logurl])
        else: 
            build.extend(['','','No log file'])
    return builds

    

def CreateExcelnew(old_suites,new_suites,builds,uuid,info,name):
    if os.path.exists(name):
        os.remove(name)
    wb = Workbook()
    ws = wb.active
    careerSheet = wb.create_sheet(0, 'career')
    url = "https://eris-portal.nvidia.com/secure/DoOneSubmissionViewCommand?osuuid=%s"%uuid
    head = ["Test Results",info[2],info[0],info[1],"%s"%url]
    print "head", head
    ws.append(head)
    head = ["Suite", "GPU", "OS", "Arch", "compiler", "chart result", "triage result","notes","log url","command"]
    ws.append(head)
    careerSheet.append("ttt")
    wb.save(name)
    exit()
    old_copy = copy.deepcopy(old_suites)
    for suite in old_copy:
        del suite[1]
        ws.append(suite)
    ws.append([])
    ws.append(["New complete"])
    new_copy = copy.deepcopy(new_suites)
    for suite in new_copy:
        del suite[1]
        ws.append(suite)
    ws.append([])
    head = ["Build Results"]
    print head
    ws.append(head)
    for build in builds:
        build[1] = ''
        ws.append(build)
    wb.save(name)

def CreateExcel(old_suites,new_suites,builds,uuid,info,name):
    if os.path.exists(name):
        os.remove(name)
    wb = Workbook()
    ws = wb.active
    url = "https://eris-portal.nvidia.com/secure/DoOneSubmissionViewCommand?osuuid=%s"%uuid
    head = ["Test Results",info[2],info[0],info[1],"%s"%url]
    print "head", head
    ws.append(head)
    head = ["Suite", "GPU", "OS", "Arch", "compiler", "chart result", "triage result","notes","log url","command"]
    ws.append(head)
    old_copy = copy.deepcopy(old_suites)
    for suite in old_copy:
        del suite[1]
        ws.append(suite)
    ws.append([])
    ws.append(["New complete"])
    new_copy = copy.deepcopy(new_suites)
    for suite in new_copy:
        del suite[1]
        ws.append(suite)
    ws.append([])
    head = ["Build Results"]
    print head
    ws.append(head)
    for build in builds:
        build[1] = ''
        ws.append(build)
    wb.save(name)


def email(user,file):
    cmd='echo "excel" | mutt -s "excel" %s@exchange.nvidia.com -a %s'%(user,file)
    os.system(cmd)
    print '&&&& Sending email to %s... Done!'%user

def SelectTestbyoption(Allstatus,Testsuites):
    Testsuites=Testsuites.split(",")
    multi_list = list()
    solo_list = list()
    for i in Testsuites:
        if i.endswith("*"):
            multi_list.append(i[:-1])
        else: solo_list.append(i)
    Selected = []
    for items in Allstatus:
        if items[0] in Testsuites: Selected.append(items)
    for items in Allstatus:
        if items not in Selected:
            for lists in multi_list:
                if items[0].startswith(lists):
                    Selected.append(items)
    return Selected

def Addfailitem(uuid,suites,subinfo,autofill):
    print 'Getting failed items...'
    SuitesListItems = list()
    if autofill:
        print "Getting autofill pre-triage for items, should take 8s for every item..."
    for suite in suites:
        Fail_list = GetFailItems(uuid,suite[0],suite[1])
        if len(Fail_list) > 5:
            Fail_list = ['MASSIVE']
        if autofill and (Fail_list[0] == 'MASSIVE' or Fail_list[0] == 'NONE'): 
            suite[8] = Addsuitehistory(uuid,suite,subinfo)
            SuitesListItems.append(suite)
            continue
        SuitesListItems.append(suite)
        for i in Fail_list:
            if 'MASSIVE' == i or 'NONE' == i:
                  break
            else:
                  if autofill:
                      if Testfailforbuild(i,uuid,suite):
                         test_result = 'Build failed'
                      else: test_result = Additemhistory(uuid,suite,i,subinfo)
                  else: test_result = ''
                  item_command = suite[10]+"_"+i.replace(' ','_')+" --test-name=\""+i+"\""
                  SuitesListItems.append(['','','','','','',i,'',test_result,'',item_command])
    return SuitesListItems

def Testfailforbuild(testname,uuid,suite):
    url = 'https://eris-portal.nvidia.com/TestResultsListDataTableServlet?&uuid1=%s&testSuiteNameText=%s&test_name_text=%s&configID=%s&command=list&attr=fail'%(uuid,suite[0],testname.replace(' ','+'),suite[1])
    try:
        url_source = GetURLSource(url)
        dict_obj = json.loads(url_source)
        assert 'Build Failed' in dict_obj["jsonArray"][0][2]
        return True
    except:
        return False
    
def GetFailItems(uuid,testsuite,configID):
    url = "https://eris-portal.nvidia.com/TestResultsListDataTableServlet?uuid1=%s&testSuiteNameText=%s&configID=%s&command=list&attr=fail"%(uuid,testsuite,configID)
    try:
        url_source = GetURLSource(url)
        dict_obj = json.loads(url_source)
        Failitems=dict_obj["funcTestFilters"]["testNames"]
        #Failitems = ['dse_dot_product.exe', 'dse_hello_events.exe', 'svm_memcpy.exe', 'svm_multi_gpu.exe', 'svm_multi_queue.exe', ...]
        if Failitems == []: Failitems = ['NONE']
        return Failitems
    except: 
        return ['NONE']

def Addsuitehistory(uuid,suite,subinfo):
    url = "https://eris-portal.nvidia.com/TestSuiteDetailsDataTableServlet?&iSortCol_0=4&sSortDir_0=desc&iDisplayStart=0&iDisplayLength=10&productNameText=%s&submTag=%s&testSuiteNameText=%s&erisConfigID=%s&command=list&attr=all"%(subinfo[0].replace(' ','+'),subinfo[1],suite[0],suite[1])
    try:
        url_source = GetURLSource(url)
        dict_obj = json.loads(url_source)
        history_suites = dict_obj["jsonArray"]
        history_len= len(history_suites)
        history_list = list()
        count = 0
        this_index = 0
        for h in history_suites:
            failcount = 0
            if 'Submission Timeout' in h[2]+h[3] or 'Failed Other' in h[2]+h[3] or 'Not Run' in h[2]+h[3] or 'Pending' in h[2]+h[3]:
                history_len = history_len-1
                continue
            if 'badge-failed' in h[14]:
                failcount = int(re.findall(r"badge-failed\'>(.*?)<",h[14])[0])
            history_list.append('massivefail' if (failcount > 4 or 'Aborted' in h[2]+h[3]) else 'notmassivefail') 
            if uuid in h[1]+h[2]: this_index = count
            count = count+1
        #history_list = ['notmassivefail','massivefail','massivefail','notmassivefail','notmassivefail',...]
        #this_index = 1
        if history_list[0] == 'notmassivefail':
            return 'No Aborted/Massive in latest %s'%subinfo[1]
        elif this_index == history_len-1:
            return 'First tested and failed Aborted/Massive'
        elif history_list[this_index+1] == 'notmassivefail':
            return 'Regression'
        else:
            fail_bias = 0
            for l in history_list[this_index+1:]:
                  fail_bias = fail_bias+1
                  if l == 'notmassivefail':
                      return 'Regression since %d %s before'%(fail_bias,subinfo[1]) 
                  elif this_index+fail_bias == history_len-1:
                      return 'Failing since first tested %d %s before'%(fail_bias,subinfo[1])
    except:
        return 'N/A'

def Additemhistory(uuid,suite,testname,subinfo):
    #uuid = '71018bcd-5309-4891-b52c-5059bda9cd2a'
    #suite = ['cuda_apps_tests_L1','231','p100','Windows10','x86_64','vc10','','','vulcan cmd']
    #testname = 'alloc64.exe'
    if not subinfo[1]== 'nightly' and not subinfo[1]== 'weekly' and not subinfo[1]== 'biweekly':
        return ''
    url = "https://eris-portal.nvidia.com/TestResultsListDataTableServlet?iDisplayStart=0&iDisplayLength=10&iSortCol_0=3&sSortDir_0=desc&productNameText=%s&submTag=%s&testSuiteNameText=%s&test_name_text=%s&configID=%s&command=list&attr=all" %(subinfo[0].replace(' ','+'),subinfo[1],suite[0],testname.replace(' ','+'),suite[1])
    try:
        url_source = GetURLSource(url)
        dict_obj = json.loads(url_source)
        history_items = dict_obj["jsonArray"]
        history_len= len(history_items)
        history_list = list()
        count = 0
        this_index = 0
        for h in history_items:
            history_list.append('pass' if 'pass' in h[2] else 'fail') #timeout regarded as fail
            if uuid in h[2]: this_index = count
            count = count+1
        #history_list = ['pass','fail','fail','pass','pass',...]
        #this_index = 1
        if history_list[0] == 'pass':
            return 'No repro in latest %s'%subinfo[1]
        elif this_index == history_len-1:
            return 'First tested and failed'
        elif history_list[this_index+1] == 'pass':
            return 'Regression'
        else:
            fail_bias = 0
            for l in history_list[this_index+1:]:
                  fail_bias = fail_bias+1
                  if l == 'pass': 
                      return 'Regression since %d %s before'%(fail_bias,subinfo[1]) 
                  elif this_index+fail_bias == history_len-1:
                      return 'Failing since first tested %d %s before'%(fail_bias,subinfo[1]) 
    except: return 'N/A'

def Addbuildtype(uuid,BuildFails,subinfo,autofill):
    if not autofill:
        return BuildFails
    print 'Getting build type,known build bug,build history...'
    #Format of Knownbugs.xlsx:  Product,Suite,GPU,OS,Arch,compiler,build mode,known bug
    if os.path.exists("Knownbugs.xlsx"):
        wb=load_workbook("Knownbugs.xlsx")
        ws=wb.active
        Bugfile_exit = True
    else: Bugfile_exit = False
    for build in BuildFails:
        url = "https://eris-portal.nvidia.com/BuildResultsListDataTableServlet?&uuid1=%s&componentNameText=%s&configID=%s&command=list&attr=fail"%(uuid,build[0],build[1])
        url_source = GetURLSource(url)
        dict_obj = json.loads(url_source)
        Buildfail=dict_obj["buildFilters"]["buildResults"]
        Buildmode=dict_obj["buildFilters"]["builds"]
        #Buildfail = ["Build Failure"]
        #Buildmode = ["release"]
        if not Buildfail[0] == 'Build Failure':
            build[7] = Buildfail[0] 
        elif Bugfile_exit:
            for line in ws:
                #build = ['cuda_apps_tests_L1','321','SLES12SP1','x86_64','','2','','','vulcan cmd']
                #line =  ['CUDA 8.0 r375','cuda_apps_tests_L1',None,'SLES12SP1','x86_64',None,'release','bug12345678']
                buildstr=build[0]+build[2]+build[3]+build[4]
                linestr=str(line[1].value)+str(line[3].value)+str(line[4].value)+str(line[5].value)
                if subinfo[0]==str(line[0].value) and Buildmode[0]==str(line[6].value) and buildstr in linestr:
                       print '***Found known build bug: ',str(line[7].value)
                       build[7]=str(line[7].value)
                       break
            if build[7] == '':
                build[7] = Addbuildhistory(uuid,build,subinfo)
    return BuildFails

def Addbuildhistory(uuid,build,subinfo):
    if not subinfo[1]== 'nightly' and not subinfo[1]== 'weekly' and not subinfo[1]== 'biweekly':
        return ''
    url = "https://eris-portal.nvidia.com/BuildResultsListDataTableServlet?iDisplayStart=0&iDisplayLength=10&iSortCol_0=5&sSortDir_0=desc&sSearch=-build&productNameText=%s&submTag=%s&componentNameText=%s&configID=%s&command=list&attr=all" %(subinfo[0].replace(' ','+'),subinfo[1],build[0],build[1])
    try:
        url_source = GetURLSource(url)
        dict_obj = json.loads(url_source)
        history_builds = dict_obj["jsonArray"]
        history_len= len(history_builds)
        history_list = list()
        count = 0
        this_index = 0
        for h in history_builds :
            history_list.append('pass' if 'pass' in h[3] else 'fail') 
            if uuid in h[2]: this_index = count
            count = count+1
        if history_list[0] == 'pass':
            return 'No repro in latest %s'%subinfo[1]
        elif this_index == history_len-1:
            return 'First built and failed'
        elif history_list[this_index+1] == 'pass':
            return 'Regression'
        else:
            fail_bias = 0
            for l in history_list[this_index+1:]:
                  fail_bias = fail_bias+1
                  if l == 'pass': 
                      return 'Regression since %d %s before'%(fail_bias,subinfo[1]) 
                  elif this_index+fail_bias == history_len-1:
                      return 'Failing since first built %d %s before'%(fail_bias,subinfo[1]) 
    except: return 'N/A'

def GetURLSource(url):
    if sys.version_info >= (2,7,9):
        ssl._create_default_https_context = ssl._create_unverified_context

    url_open = urllib2.urlopen(url)
    url_source = url_open.read()
    return url_source


def main():
    parser = optparse.OptionParser()
    parser.add_option('-u','--uuid',dest='uuid',action='store',type='string',default=False,help='specify the uuid of eris nightly')
    parser.add_option('-e','--email',dest='email',action='store',type='string',default=False,help='specify the mail username if needed')
    parser.add_option('-t','--testsuites',dest='testsuites',action='store',type='string',default=False,help='specify testsuites/componets,for example:--testsuites=npp*,cublas*,gdk_tests')
    parser.add_option('-d','--detail',dest='detail',action='store_true',default=False,help='Get and record fail items')
    parser.add_option('-a','--autofill',dest='autofill',action='store_true',default=False,help='Autofill pre-triage result for every build and test item')
    options,args = parser.parse_args()
    if options.uuid:
        subinfo = GetSubInfo(options.uuid)
        #subinfo = ('CUDA 8.0 r375','nightly','01-Oct-2016','//sw/rel/gpgpu/toolkit/r8.0/eris/cuda_r375.vlcp')
        excelname = os.path.join(sys.path[0],subinfo[0].replace(' ','_')+'_'+subinfo[1]+'_'+subinfo[2].replace('-','_')+'.xlsx')
        #excelname = os.path.join(sys.path[0],"Chart_result.xlsx")
        if options.testsuites:
            Testsuites=options.testsuites
        else:
            Testsuites = list()
            txt = os.path.join(sys.path[0],"Create_excel_components.txt")
            with open(txt,'r') as file:
                for line in file.readlines():
                    if not line.startswith('#') and not line == "\n" and not line=="\r":
                        component = line.strip('\n').strip('\r')
                        Testsuites.append(component)

        Allstatus = GetRunConfigs(options.uuid)
        FailsuitesList = SelectFailTestsuites(Allstatus,Testsuites,options.testsuites)
        oldsuitesList,newsuitesList = Selectnewcomplete(FailsuitesList,options.uuid)
        for i in oldsuitesList:
            print i
        print "============O"
        for i in newsuitesList:
            print i
        Allbuild = Getbuildresults(options.uuid,Testsuites)
        BuildFailList = SelectFailTestsuites(Allbuild,Testsuites,options.testsuites)
        print "============B"
        for i in  BuildFailList:
            print i
        print "============L"
        print "Getting log urls and commands..."
        oldsuitesList_log = Addtestlog(options.uuid,oldsuitesList,subinfo)
        newsuitesList_log = Addtestlog(options.uuid,newsuitesList,subinfo)
        BuildFailList_log = Addbuildlog(options.uuid,BuildFailList,subinfo)
        oldsuitesList_cmd = Addtestcmd(options.uuid,oldsuitesList_log,subinfo[3])
        newsuitesList_cmd = Addtestcmd(options.uuid,newsuitesList_log,subinfo[3])
        BuildFailList_cmd = Addbuildcmd(options.uuid,BuildFailList_log,subinfo[3])
        if options.detail:
            oldsuitesList_item = Addfailitem(options.uuid,oldsuitesList_cmd,subinfo,options.autofill)
            newsuitesList_item = Addfailitem(options.uuid,newsuitesList_cmd,subinfo,options.autofill)
            BuildFailList_cmd_tpye = Addbuildtype(options.uuid,BuildFailList_cmd,subinfo,options.autofill)
            #CreateExcel(oldsuitesList_item,newsuitesList_item,BuildFailList_cmd_tpye,options.uuid,subinfo,excelname)
        else:
            #CreateExcel(oldsuitesList_cmd,newsuitesList_cmd,BuildFailList_cmd,options.uuid,subinfo,excelname)            
        
        if options.email:
                email(options.email,excelname)

    else:
        print 'ERROR: This script need option "--uuid" or "-u".'
        sys.exit(1) 

if __name__=="__main__":
        main()
        
        

