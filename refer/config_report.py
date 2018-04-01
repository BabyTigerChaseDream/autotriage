import os
import sys
import time
import urllib2
import urllib
import re
import optparse
import json
try:
        from openpyxl import Workbook
except:
        print "\n[Error:]You need to install \"openpyxl\" to execute the program\n"
        print "  On Linux - Ubuntu:"
        print "  $ sudo apt-get install python-openpyxl\n"
        print "  On Windows :"
        print "  Please follow the blog: http://blog.csdn.net/a102111/article/details/46785745"
        sys.exit()
from openpyxl import load_workbook
from openpyxl.styles import Font
import  openpyxl.styles.colors as Color
import shutil

def Get_param():
	parser=optparse.OptionParser()
	parser.add_option("-t", "--template",dest="template",help="the template path, for example: ./GP102_template.xlsx")
	(options,args)=parser.parse_args()
	return options
                 
def GetProducts(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        print "********Products******"
        products = []
        productflg = False
        for i in range(1,300):
                if productflg == False:
                        if not ws["A%s"%i].value == None:
                                if 'Branch' in ws["A%s"%i].value:
                                        productflg = True
                        continue
                if ws["A%s"%i].value==None and ws["B%s"%i].value==None:
                        if not products == []:
                                products[len(products)-1].append(i)
                        break
                if not ws["A%s"%i].value==None:
                        productinfo = []
                        productinfo.append(ws["A%s"%i].value)
                        productinfo.append(i)
                        if not products == []:
                                products[len(products)-1].append(i)
                        products.append(productinfo)
        for p in products:
                print p[0].strip()
                     
        configs=[]
        for m in range(99,123):
                mc = chr(m)
                if  ws["%s3"%mc] == None:
                        break
                if not ws["%s1"%mc].value == None:
                        config_str = ws["%s1"%mc].value + "+" + ws["%s2"%mc].value
                        configs.append(config_str)
        configlist=list()
        for config in configs:
                if 'x86_64' in config.lower():
                        config = [config]
                        config.append('r375')
                        configlist.append(config)
                        continue
                elif 'ppc64le' in config.lower():
                        config = [config]
                        config.append('r375')
                        configlist.append(config)
                        continue
                       
        print "********Configs******"
        return products, configlist

def GetTopuuid(product,branch):
        print 'UUID: ',
        
        if  branch == 'r361' :
                ProductUrls = {
                "CUDA r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDA+8.0+r361&paramListSubmissionsTagFilter=nightly",
                "cuDNN v6 r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDNN+V6+CUDA+8.0+r361&paramListSubmissionsTagFilter=nightly",
                "NCCL public r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=NCCL+Public+V1.2+CUDA+8.0+r361&paramListSubmissionsTagFilter=nightly",
                "OpenCL (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=OpenCL+gpgpu+r361&paramListSubmissionsTagFilter=nightly",
                "OpenMPI r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=OpenMPI+CUDA+8.0+r361&paramListSubmissionsTagFilter=nightly",
                "CUDA r8.0 (weekly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDA+8.0+r361&paramListSubmissionsTagFilter=weekly",
                "cuDNN v6 r8.0 (weekly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDNN+V6+CUDA+8.0+r361&paramListSubmissionsTagFilter=weekly",
                "OpenCL (weekly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=OpenCL+gpgpu+r361&paramListSubmissionsTagFilter=weekly"
                }
                
        if  branch == 'r367' :
                ProductUrls = {
                "CUDA r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDA+8.0+r367&paramListSubmissionsTagFilter=nightly",
                "cuDNN v6 r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDNN+V6+CUDA+8.0+r367&paramListSubmissionsTagFilter=nightly",
                "NCCL public r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=NCCL+Public+V1.2+CUDA+8.0+r367&paramListSubmissionsTagFilter=nightly",
                "OpenCL (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=OpenCL+gpgpu+r367&paramListSubmissionsTagFilter=nightly",
                "OpenMPI r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=OpenMPI+CUDA+8.0+r367&paramListSubmissionsTagFilter=nightly",
                "CUDA r8.0 (weekly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDA+8.0+r367&paramListSubmissionsTagFilter=weekly",
                "OpenCL (weekly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=OpenCL+gpgpu+r367&paramListSubmissionsTagFilter=weekly"
                }                        
                
        if  branch == 'r375' :
                ProductUrls = {
                "CUDA r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDA+8.0+r375&paramListSubmissionsTagFilter=nightly",
                "cuDNN v6 r8.0 (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDNN+V6+CUDA+8.0+r375&paramListSubmissionsTagFilter=nightly",
                "OpenCL (nightly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=OpenCL+gpgpu+r375&paramListSubmissionsTagFilter=nightly",
                "CUDA r8.0 (weekly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDA+8.0+r375&paramListSubmissionsTagFilter=weekly",
                "cuDNN v6 r8.0 (weekly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=CUDNN+V6+CUDA+8.0+r375&paramListSubmissionsTagFilter=weekly",
                "OpenCL (weekly)" : "https://eris-portal.nvidia.com/ListSubmissionsServlet?&sSearch=sw-cuda-qa-sh&iSortCol_0=2&sSortDir_0=desc&productNameText=OpenCL+gpgpu+r375&paramListSubmissionsTagFilter=weekly"
                }
                
        url = ''
        for key in ProductUrls:
                if key.lower() in product.lower():
                        url = ProductUrls[key]
                        break
        if url == '':
                print 'None'
                return None,None
        else:
                json_str = urllib2.urlopen(url).read()
                dict_obj = json.loads(json_str)
                uuidstr = dict_obj["jsonArray"][0][1]
                uuidtime = dict_obj["jsonArray"][0][2]
                uuidlist = re.findall("osuuid\=(.*?)\'",uuidstr)
        print uuidlist[0]
        return uuidlist[0],uuidtime


def GetRunConfigs(uuid,configs_id):
     #configs_id = [['Ubuntu16_04/x86_64+GP102(TITAN X)', '4147'], ['windows10/x86_64+GP102(TITAN X)', '17584']]
     print 'Results: ',
     header_url="https://eris-portal.nvidia.com/GetSubmissionHeaderTwoUuidsServlet?uuid1=%s"%uuid
     header_results = urllib2.urlopen(header_url).read()
     header = re.findall("Status.*?<td>(.*?)<",header_results)
     print header[0]
     eris_url = "https://eris-portal.nvidia.com/OneSubmissionTestSuitesServlet?uuid=%s&tableselector=dataTableOneSubmissionTestSuites&url"%uuid
     eris_handle = urllib2.urlopen(eris_url)
     eris_results = eris_handle.read()
     Reportdict = dict()

     for config in configs_id:    
             PassConfigsTp= re.findall(r"badge-(passed)\'>.{1,6}<\\/span><\\/a> 0 0<.{1,80}historyChannelLabel.{1,180}testSuiteNameText=(.{1,40})\&erisConfigID\=%s\'"%config[1],eris_results)
             #PassConfigsTp=[('passed','cuda_math_sm2xplus_tests_L0'),('passed','cuda_math_sm2xplus_tests_L1')]
             tempdictp = dict()
             for i in PassConfigsTp:
                     tempdictp[i[1]] = 'Pass'
             #tempdictp = {"cuda_apps_tests_L0":"Pass", "cuda_apps_tests_L1":"Pass"}

             PassConfigsTp2= re.findall(r"badge-(passed)\'>.{1,6}<\\/span><\\/a> 0 <a id=\'&testSuiteNameText=(.{1,40})%s\'"%config[1],eris_results)
             #PassConfigsTp2=[('passed','cuda_math_sm2xplus_tests_L0'),('passed','cuda_math_sm2xplus_tests_L1')]
             tempdictp2 = dict()
             for i in PassConfigsTp2:
                     tempdictp2[i[1]] = 'Pass'
             #tempdictp2 = {"cuda_apps_tests_L0":"Pass", "cuda_apps_tests_L1":"Pass"}

             FailConfigs= re.findall(r"badge-(failed)\'>.{1,6}<\\\/span><\\\/a> 0<.{1,100}historyChannelLabel.{1,180}testSuiteNameText=(.{1,40})&erisConfigID=%s\'>"%config[1],eris_results)
             #FailConfigs=[('failed','cublas_cnp_tests_L0'), ('failed','cublas_cnp_tests_L1')]
             tempdictf = dict()
             for i in FailConfigs:
                     tempdictf[i[1]] = 'Fail'
             #tempdictf = {"cuda_apps_tests_L0":"Fail", "cuda_apps_tests_L1":"Fail"}
                     
             FailConfigs2= re.findall(r"badge-(failed)\'>.{1,6}<\\\/span><\\\/a> <a id=\'&testSuiteNameText=(.{1,40})%s\'"%config[1],eris_results)
             #FailConfigs2=[('failed','cublas_cnp_tests_L0'), ('failed','cublas_cnp_tests_L1')]
             tempdictf2 = dict()
             for i in FailConfigs2:
                     tempdictf2[i[1]] = 'Fail'
             #tempdictf2 = {"cuda_apps_tests_L0":"Fail", "cuda_apps_tests_L1":"Fail"}

             AbortConfigs= re.findall(r"badge-aborted\'>(Aborted)<.{1,100}historyChannelLabel.{1,180}testSuiteNameText=(.{1,40})&erisConfigID=%s\'>"%config[1],eris_results)
             #AbortConfigs=[('failed','cublas_cnp_tests_L0'), ('failed','cublas_cnp_tests_L1')]
             tempdicta = dict()
             for i in AbortConfigs:
                     tempdicta[i[1]] = 'Aborted'
             #tempdicta = {"cuda_apps_tests_L0":"Aborted", "cuda_apps_tests_L1":"Aborted"}
            
             PendingConfigs= re.findall(r"transparentTable\'>(Pending)<.{1,80}historyChannelLabel.{1,180}testSuiteNameText=(.{1,40})&erisConfigID=%s\'"%config[1],eris_results)
             #PendingConfigs = [('Pending','cuda_apps_tests_L0'),('Pending','cuda_apps_tests_L1')]
             tempdictm = dict()
             for i in PendingConfigs:
                     tempdictm[i[1]] = 'Pending'
             #tempdictm = {"cuda_apps_tests_L0":"not run", "cuda_apps_tests_L1":"not run"}

             RunningConfigs= re.findall(r"transparentTable\'>(Running)<.{1,80}historyChannelLabel.{1,180}testSuiteNameText=(.{1,40})&erisConfigID=%s\'"%config[1],eris_results)
             #RunningConfigs = [('Running','cuda_apps_tests_L0'),('Running','cuda_apps_tests_L1')]
             tempdictr = dict()
             for i in RunningConfigs:
                     tempdictr[i[1]] = 'running'
             #tempdictr = {"cuda_apps_tests_L0":"running", "cuda_apps_tests_L1":"running"}

             
             FailOtherConfigs= re.findall(r"badge-failed\'>(Failed Other)<.{1,100}historyChannelLabel.{1,180}testSuiteNameText=(.{1,40})&erisConfigID=%s\'"%config[1],eris_results)
             #FailOtherConfigs=[('Failed Other','cublas_cnp_tests_L0'), ('Failed Other','cublas_cnp_tests_L1')]
             tempdicto = dict()
             for i in FailOtherConfigs:
                     tempdicto[i[1]] = 'Failed Other'
             #tempdict0 = {"cuda_apps_tests_L0":"Failed Other", "cuda_apps_tests_L1":"Failed Other"}
                           
             Reportdict[config[0]] = dict(tempdictp.items()+tempdictp2.items()+tempdictf.items()+tempdictf2.items()+tempdicta.items()+tempdictm.items()+tempdictr.items()+tempdicto.items())
             #Reportdict = {"Windows10/x86_64+GP106(gtx1060)": {"cuda_apps_tests_L0":"Pending", "cuda_apps_tests_L1":"Pending","cuda_math_tests_L0":"failed"}}
     #Reportdict = {"Windows10/x86_64+GP106(gtx1060)": {"cuda_apps_tests_L0":"Pending", "cuda_apps_tests_L1":"Pending","cuda_math_tests_L0":"Fail"}, "Ubuntu16_04/x86_64+GP102(TITAN X)": {"aaa" : "Pass","bbb": "Aborted"}}
     for i in Reportdict:
             print '[',i ,']', ' : '
             for suite in Reportdict[i]:
                     print suite, ' : ',  Reportdict[i][suite]
     return Reportdict


def UpdateExcel(filepath,product,subuuid,Tsub,report,count):
        eris_link = "https://eris-portal.nvidia.com/secure/DoOneSubmissionViewCommand?osuuid=%s"%subuuid
        wb = load_workbook(filepath)
        ws = wb.active
        print 'Updating Excel...',
        try:
                for m in range(99,123):
                        mc = chr(m)
                        if ws['%s3'%mc].value == None:
                                ws['%s%s'%(chr(m+1),product[1]+count)] = Tsub
                                ws['%s%s'%(chr(m+2),product[1]+count)] = eris_link
                                break
                        if not ws['%s3'%mc].value == None and 'New bugs' in ws['%s3'%mc].value:
                                for config in report:
                                        if ws['%s1'%mc].value.lower() == config.split('+')[0].lower() and ws['%s2'%mc].value.lower() == config.split('+')[1].lower():
                                                for i in range(int(product[1]),int(product[2])):
                                                        if ws['%s%s'%(mc,i)].value == 'N/A' :
                                                                continue
                                                        try:
                                                                ws['%s%s'%(mc,i)] = report['%s'%config]['%s'%(ws['B%s'%i].value)]
                                                        except:
                                                                ws['%s%s'%(mc,i)] = 'not run'
                                                                pass
                                                        if ws['%s%s'%(mc,i)].value == "Failed Other" and 'ppc64le' not in config:
                                                                ws['%s%s'%(mc,i)] = 'not run'
                                                        if ws['%s%s'%(mc,i)].value == "Pass":
                                                                ws['%s%s'%(mc,i)].font = Font(color="00CD00") # Dark Grean
                                                        if ws['%s%s'%(mc,i)].value == "Fail" or ws['%s%s'%(mc,i)].value == "Aborted" or ws['%s%s'%(mc,i)].value == "Failed Other":
                                                                ws['%s%s'%(mc,i)].font = Font(color="FF0000") # Red
                                                        if ws['%s%s'%(mc,i)].value == "running":
                                                                ws['%s%s'%(mc,i)].font = Font(color="FF00FF") # Purple
                                                        if ws['%s%s'%(mc,i)].value == "not run" or ws['%s%s'%(mc,i)].value == "Pending":
                                                                ws['%s%s'%(mc,i)].font = Font(color="000000") # Black
                                                        
                
        finally:
                wb.save(filepath)
                print "Done"


def GetConfigID(configdict):
        newconfigdict = dict()
        for branch in configdict:
                configsuite = list()
                for config in configdict[branch]:
                        configtemp = list()
                        configtemp.append(config)
                        configID = ''
                        if 'gp102' in config.lower() and 'x86_64' in config.lower() and 'windows10' in config.lower():
                                configID = '17584'
                        if 'gp102' in config.lower() and 'x86_64' in config.lower() and 'ubuntu16_04' in config.lower():
                                configID = '4147'
                        if 'p100' in config.lower() and 'x86_64' in config.lower() and 'windows10' in config.lower():
                                configID = '10572'
                        if 'gp100sxm2' in config.lower() and 'ppc64le' in config.lower() and 'rhel7_3' in config.lower():       
                                configID = '288402'
                        if 'gp100sxm2' in config.lower() and 'ppc64le' in config.lower() and 'ubuntu16_04' in config.lower():       
                                configID = '291419'
                        if 'p100' in config.lower() and 'x86_64' in config.lower() and 'rhel7_3' in config.lower():
                                configID = '2527245'
                                if 'clang-llvm-3.8.1' in config.lower():
                                        configID = '2540830'
                                elif 'pgi16_7' in config.lower():
                                        configID = '2540658'

                                
                        configtemp.append(configID)
                        configsuite.append(configtemp)
                newconfigdict[branch] = configsuite
        return newconfigdict

def  ConfigSort(configs):
        configdict= dict()
        configlist = list()
        for config in configs:
                configlist = []
                if not configdict.has_key(config[1]):
                        configdict[config[1]]= [config[0]]
                else:
                        configlist = configdict[config[1]]
                        configlist.append(config[0])
                        configdict[config[1]]= configlist
        for branch in configdict:
                print branch, ':'
                for config in configdict[branch]:
                        print config
        return configdict
        
        
def main():

        #Solve the ssl failure when Python > 2.7.9
        import ssl 
        ssl._create_default_https_context = ssl._create_unverified_context
        
        options = Get_param()
        filepath = options.template
        if 'gp102' in filepath.lower():
                reportfile = 'GP102_Report.xlsx'
        if 'gp100' in filepath.lower(): 
                reportfile = 'GP100_Report.xlsx'
        shutil.copy(filepath,reportfile)
        (products, configs) = GetProducts(reportfile)
        #products = [["CUDA r8.0 (nightly)",4,53] ["cuDNN v5.1 r8.0 (nightly)",55,76]]
        #configs =  [["Ubuntu16_04/x86_64+GP102(TITAN X)",'r367'],["Windows10/x86_64+GP106(gtx1060)",'r375']]
        configdict = ConfigSort(configs)
        #configdict = {'r367': ['Ubuntu16_04/x86_64+GP102(TITAN X)', 'RHEL7_2/x86_64+GP104(TITAN X)'], 'r375': ['Windows10/x86_64+GP106(gtx1060)', 'Windows10/x86_64+GP100(p100)'], 'r361': ['Ubuntu14_04/x86_64+gm204(m60)']}
        newconfigdict = GetConfigID(configdict)
        #newconfigdict = {'r367': [['Ubuntu16_04/x86_64+GP102(TITAN X)', '4147'], ['Windows10/x86_64+GP102(TITAN X)', '17584']], 'r375': [['Windows10/x86_64+p100', '10572'], ['RHEL7_3/ppc64le+GP100(p100)', '288402']]}
        for product in products:
                count = 0
                print '\n\n=============', product[0].replace('\n',''),'============='
                for branch in newconfigdict:
                        Uuid,Tuuid = GetTopuuid(product[0],branch)
                        if Uuid == None:
                                continue
                        Report = GetRunConfigs(Uuid,newconfigdict[branch])
                        #Report = {"Windows10/x86_64+GP106(gtx1060)": {"cuda_apps_tests_L0":"Pending", "cuda_apps_tests_L1":"Pending","cuda_math_tests_L0":"Fail"},"Ubuntu16_04/x86_64+GP102(TITAN X)": {"aaa" : "Pass","bbb": "Aborted"}}
                        UpdateExcel(reportfile,product,Uuid,Tuuid,Report,count)
                        count = count + 1
                print '=============================================================='
        print "\n\n\nOutput as [", reportfile, "]"


if __name__ == "__main__":
         main()

    
